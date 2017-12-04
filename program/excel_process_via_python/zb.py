import math
import openpyxl
import numpy as np

wb = openpyxl.load_workbook('ZB.xlsx')
sheet = wb.get_sheet_by_name('青藏_16S')
sheet_2 = wb.get_sheet_by_name('内蒙_16S')
wb.create_sheet('combined')
wb.save('ZB.xlsx')
sheet_combine = wb.get_sheet_by_name('combined')

def read_excel(sheet_name, row, col):
    id = []
    data = []
    name = []
    # read id
    for x in range(1, row+1):
        for y in range(1, 2):
            id.append(sheet_name.cell(row=x,column=y).value)

    for x in range(1, 2):
        for y in range(1, col+1):
            name.append(sheet_name.cell(row=x,column=y).value)

    for x in range(1, row+1):
        temp = []
        for y in range(1, col+1):
            temp.append(sheet_name.cell(row=x,column=y).value)
        data.append(temp)

    return id, name, data

# combine two lines and write to table
def c2w(line_number, line1, line2, sheet_name):
    new_line = line1 + line2[1:]
    for i in range(0, len(new_line)):
        sheet_name.cell(row=line_number, column=i+1).value = new_line[i]

# print(read_excel(sheet_2,9552,163))
id_1, name_1, data_1 = read_excel(sheet,10686, 129)
id_2, name_2, data_2= read_excel(sheet_2, 9552, 163)

current_line = 1
c2w(current_line, name_1, name_2, sheet_combine)

for i in range(1,len(id_1)):
    current_line += 1
    current_id = id_1[i]
    for j in range(1, len(id_2)):
        if current_id == id_2[j]:
            print("id equas")
            c2w(current_line, data_1[i], data_2[j], sheet_combine)# write two line in table 3
            break
    else:
        c2w(current_line, data_1[i], [], sheet_combine) # only write table 1 in table 3

for i in range(0, len(id_2)):
    current_id = id_2[i]
    if current_id in id_1:
        pass
    else:
        current_line += 1
        blank = []
        blank.append(current_id)
        for j in range(1, len(name_1)):
            blank.append(" ")
        c2w(current_line, blank, data_2[i], sheet_combine) # write line in table 3

print(name_1)
print(name_2)

print('*'*80)
print(id_1)
print(id_2)
print('*'*80)
for i in range(0, 4):
    print(data_1[i])

wb.save('ZB.xlsx')