import math
import openpyxl
import numpy as np

def read_excel(sheet_name, row, col):
    id = [] # all the first column info
    name = [] # all the first row info
    data = [] # all data by rows

    for x in range(1, row+1):
        for y in range(1, 2):
            id.append(sheet_name.cell(row=x,column=y).value)

    for x in range(1, 2):
        for y in range(1, col+1):
            name.append(sheet_name.cell(row=x,column=y).value)

    for x in range(1, row+1):
        temp = []
        for y in range(1, col+1):
            temp.append(sheet_name.cell(row=x,column=y).value)
        data.append(temp)

    return id, name, data

# combine two lines and write to table
def c2w(line_number, line1, line2, sheet_name):
    new_line = line1 + line2[1:]
    for i in range(0, len(new_line)):
        sheet_name.cell(row=line_number, column=i+1).value = new_line[i]

def c2w_head(line_number, line1, line2, sheet_name):
    head_1 = [line1[0]] + [sheet_name_1+'.'+str(i) for i in line1[1:]]
    print('x',head_1)
    head_2 = [sheet_name_2+'.'+str(i) for i in line2[1:]]
    print('x',head_2)
    new_line = head_1 + head_2
    print(new_line)
    for i in range(0, len(new_line)):
        sheet_name.cell(row=line_number, column=i+1).value = new_line[i]

row_1, col_1 = 48756, 129
row_2, col_2 = 24012, 163

sheet_name_1 = '5'
sheet_name_2 = '6'
# file_name = 'Yangdai_otutable_unoise3_minsize2.xlsx'
file_name = 'Yangdai_otutable_unoise3_minsize2.xlsx'
wb = openpyxl.load_workbook(file_name)

print('load finished')
sheet_1 = wb.get_sheet_by_name(sheet_name_1)
sheet_2 = wb.get_sheet_by_name(sheet_name_2)

id_1, name_1, data_1 = read_excel(sheet_1, row_1, col_1)
id_2, name_2, data_2 = read_excel(sheet_2, row_2, col_2)

print('start to process')
current_line = 1

file_combine_name = 'data_5_6.xlsx'
wb_combine_aim = openpyxl.Workbook()
wb_combine_aim.save(file_combine_name)
wb_combine = openpyxl.load_workbook(file_combine_name)
first_sheet = wb_combine.worksheets[0]
first_sheet.title = 'combined_5_6'
sheet_combine = wb_combine.get_sheet_by_name('combined_5_6')

c2w_head(current_line, name_1, name_2, sheet_combine) # write the head of excel

zeros_1 = [0 for i in range(1, len(name_1))]
zeros_2 = [0 for i in range(1, len(name_2)+1)]

for i in range(1,len(id_1)):
    current_line += 1
    current_id = id_1[i]
    for j in range(1, len(id_2)):
        if current_id == id_2[j]:
            print(i,":id equas")
            c2w(current_line, data_1[i], data_2[j], sheet_combine)# write two line in table 3
            break
    else:
        c2w(current_line, data_1[i], zeros_2, sheet_combine) # only write table 1 in table 3

print('here')
for i in range(0, len(id_2)):
    print(i, ':')
    current_id = id_2[i]
    if current_id in id_1:
        pass
    else:
        current_line += 1
        blank = []
        blank.append(current_id)
        blank = blank + zeros_1
        c2w(current_line, blank, data_2[i], sheet_combine) # write line in table 3
print('write to new file')
wb_combine.save(file_combine_name)